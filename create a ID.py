#!/usr/bin/env python
# coding: utf-8

# In[4]:


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from datetime import datetime

# Function to prompt for class ID counts
def get_class_id_counts():
    classes = [
        "std-1", "std-2", "std-3", "std-4", "std-5", "std-6",
        "std-7", "std-8", "std-9", "std-10", "std-11S", "std-11C", "std-11A", "std-12S", "std-12C", "std-12A"
    ]
    id_counts = {}
    for class_name in classes:
        try:
            count = int(input(f"Enter ID count for {class_name}: "))
            id_counts[class_name] = count
        except ValueError:
            print(f"Invalid input for {class_name}, defaulting to 0.")
            id_counts[class_name] = 0
    return id_counts

# Function to process student class based on board and language rules
def process_student_class(student_class_number, board, language):
    if board == "cbse":
        if 1 <= student_class_number <= 2:
            return student_class_number
        elif 3 <= student_class_number <= 12:
            return student_class_number + 1
        elif student_class_number in ["11A", "11C", "11S", "12A", "12C", "12S", "Nursery"]:
            return "46"
        elif student_class_number == "Junior KG":
            return "47"
        elif student_class_number == "Senior KG":
            return "48"
        else:
            return "class not match"
    elif board == "gseb":
        if language == "en":
            if 1 <= student_class_number <= 12:
                return student_class_number + 25
            else:
                return "class not match"
        elif language == "gu":
            if 1 <= student_class_number <= 10:
                return student_class_number + 13
            elif student_class_number == "11A":
                return "41"
            elif student_class_number == "11C":
                return "38"
            elif student_class_number == "11S":
                return "24"
            elif student_class_number == "12A":
                return "42"
            elif student_class_number == "12C":
                return "39"
            elif student_class_number == "12S":
                return "25"
            else:
                return "class not match"
    return "class not match"

# Function to format and validate email
def format_email(email):
    # Trim whitespace and convert to lowercase
    return email.strip().lower()

# Step 1: Collecting inputs
short_key = input("Enter Short Key: ")
school_name = input("Enter School Name: ")
board = input("Enter Board (cbse/gseb): ").lower()
language = input("Enter Language (en/gu): ").lower()
district = input("Enter District: ")
institute_id = input("Enter Institute ID: ")
user_type = input("Enter User Type: ")

# Collect ID counts dynamically
id_counts = get_class_id_counts()

# Step 2: Generate the data
records = []
additional_sheet_records = []
for class_name, id_count in id_counts.items():
    student_class_number = int(class_name.split('-')[1]) if class_name.split('-')[1].isdigit() else 0
    # student_original_class =  int(class_name.split('-')[1]) if class_name.split('-')[1].isdigit() else 0
    for i in range(1, id_count + 1):
        student_name = f'student{i}'
        student_email = format_email(f'{student_name}@{short_key}.melzo')
        
        # Use the student class number derived from the class name
        processed_class = process_student_class(student_class_number, board, language)
        
        student_row = {
            'Sr No': i,
            'STUDENTNAME': student_name,
            'STUDENTMAIL': student_email,
            'STUDENTMOB': '1234567890',
            'studentuserid': student_email,
            'originalClasss': student_class_number,
            'studentclass': processed_class,  # Updated processed class
            'institute': institute_id,
            'schoolname': school_name,
            'board': board,
            'lang': language,
            'district': district,
            'usertype': user_type
        }
        additional_row = {
            'Sr No': i,
            'STUDENTNAME': student_name,
            'User-Id': student_email,
            'Class': student_class_number,
            'Div': 'A',
            'Password': '12345678'
        }
        records.append(student_row)
        additional_sheet_records.append(additional_row)

# Step 3: Create DataFrames
df = pd.DataFrame(records)
additional_df = pd.DataFrame(additional_sheet_records)

# Step 4: Generate Excel file name
filename = f"{school_name}_{district}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


# 

# Step 3: Segregate the additional records based on class
segregated_data = {}
for row in additional_sheet_records:
    class_value = row['Class']
    if class_value not in segregated_data:
        segregated_data[class_value] = []
    segregated_data[class_value].append(row)

# # Step 4: Create a new Excel file with a sheet for each class
# filename = f"{school_name}_{district}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
# with pd.ExcelWriter(filename, engine='openpyxl') as writer:
#     # Create separate sheets for each class in the segregated data
#     for class_value, class_records in segregated_data.items():
#         df_class = pd.DataFrame(class_records)
#         sheet_name = f"Class_{class_value}"
#         df_class.to_excel(writer, sheet_name=sheet_name, index=False)

# 

# Step 5: Save the Excel file with all sheets
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Student_Form', index=False)
    additional_df.to_excel(writer, sheet_name='AdditionalSheet', index=False)
    for class_value, class_records in segregated_data.items():
        df_class = pd.DataFrame(class_records)
        sheet_name = f"Class_{class_value}"
        df_class.to_excel(writer, sheet_name=sheet_name, index=False)

# Step 6: Apply formatting to all sheets with openpyxl
wb = load_workbook(filename)

# Define cell borders and styles
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='fbbc04', end_color='fbbc04', fill_type='solid')
bold_font = Font(bold=True)

# Loop through all sheets in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Apply formatting to the header row (1st row)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font

    # Apply borders and bold font to non-empty cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:  # Apply borders and bold if the cell is not empty
                cell.font = bold_font
                cell.border = thin_border

# Save the updated Excel file
wb.save(filename)

print(f"Excel file '{filename}' generated successfully with formatted sheets!")

